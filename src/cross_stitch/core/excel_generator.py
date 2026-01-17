"""Excel generation functionality for cross-stitch patterns."""

from typing import List, Dict, Any, Union
from pathlib import Path
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import PatternSet, CrossStitchPattern, GeneratorConfig
from ..utils import ExcelGenerationError, save_file


class ExcelGenerator:
    """Generates Excel files from cross-stitch patterns."""

    def __init__(self, config: GeneratorConfig):
        """
        Initialize ExcelGenerator with configuration.

        Args:
            config: Generator configuration object
        """
        self.config = config

    def generate_excel_file(
        self, pattern_set: PatternSet, output_path: Union[str, Path]
    ) -> Path:
        """
        Generate complete Excel file from pattern set.

        Args:
            pattern_set: Collection of cross-stitch patterns
            output_path: Path where to save the Excel file

        Returns:
            Path to the generated Excel file

        Raises:
            ExcelGenerationError: If Excel generation fails
        """
        try:
            # Create new workbook
            workbook = Workbook()

            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            # Create pattern sheets for each resolution
            for resolution_name in pattern_set.resolution_names:
                pattern = pattern_set.get_pattern(resolution_name)
                self._create_pattern_sheet(workbook, pattern)

            # Create color legend sheet if requested
            if self.config.include_color_legend:
                self._create_legend_sheet(workbook, pattern_set)

            # Create summary sheet with metadata
            self._create_summary_sheet(workbook, pattern_set)

            # Save workbook to bytes
            excel_bytes = self._save_workbook_to_bytes(workbook)

            # Save to file
            output_file_path = save_file(excel_bytes, output_path)

            return output_file_path

        except Exception as e:
            if isinstance(e, ExcelGenerationError):
                raise
            raise ExcelGenerationError(
                f"Failed to generate Excel file: {e}",
                output_path=str(output_path),
                cause=e,
            )

    def _create_pattern_sheet(
        self, workbook: Workbook, pattern: CrossStitchPattern
    ) -> None:
        """
        Create a worksheet for a specific pattern resolution.

        Args:
            workbook: Excel workbook object
            pattern: Cross-stitch pattern to generate sheet for
        """
        try:
            # Create worksheet with resolution name
            sheet_name = self._sanitize_sheet_name(pattern.resolution_name)
            worksheet = workbook.create_sheet(title=sheet_name)

            # Set up the worksheet structure
            self._setup_pattern_worksheet(worksheet, pattern)

            # Apply pattern colors
            self._apply_pattern_colors(worksheet, pattern)

            # Apply cell styling
            self._apply_cell_styling(worksheet, pattern)

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to create pattern sheet: {e}",
                sheet_name=pattern.resolution_name,
                cause=e,
            )

    def _setup_pattern_worksheet(
        self, worksheet: Worksheet, pattern: CrossStitchPattern
    ) -> None:
        """Set up basic structure of pattern worksheet with professional rulers."""
        # Set cell dimensions to make them square
        cell_size_points = self.config.excel_cell_size

        # Convert points to Excel column width units for square cells
        # Excel column width is in character units (based on default font)
        # More accurate conversion: points / 7.5 works better for square cells
        excel_width = cell_size_points / 7.5

        # Pattern is offset by 1 row and 1 column for rulers
        # Column A is reserved for row rulers, Row 1 for column rulers
        for col in range(1, pattern.width + 2):  # +2 for ruler column
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = excel_width

        # Set row heights (in points) - +2 for ruler row
        for row in range(1, pattern.height + 2):
            worksheet.row_dimensions[row].height = cell_size_points

        # Add professional rulers
        self._add_custom_rulers(worksheet, pattern)

    def _add_custom_rulers(self, worksheet: Worksheet, pattern: CrossStitchPattern) -> None:
        """Add custom ruler system for professional cross-stitch counting."""
        # Ruler styling
        ruler_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        ruler_font = Font(bold=True, size=10)
        ruler_alignment = Alignment(horizontal="center", vertical="center")

        # Column rulers (Row 1): Show stitch numbers every 10 stitches
        for col in range(10, pattern.width + 1, 10):
            excel_col = col + 1  # +1 for offset (ruler column A)
            cell = worksheet.cell(row=1, column=excel_col, value=col)
            cell.fill = ruler_fill
            cell.font = ruler_font
            cell.alignment = ruler_alignment

        # Row rulers (Column A): Show row numbers every 10 rows
        for row in range(10, pattern.height + 1, 10):
            excel_row = row + 1  # +1 for offset (ruler row 1)
            cell = worksheet.cell(row=excel_row, column=1, value=row)
            cell.fill = ruler_fill
            cell.font = ruler_font
            cell.alignment = ruler_alignment

        # Leave A1 blank as recommended
        a1_cell = worksheet.cell(row=1, column=1)
        a1_cell.fill = ruler_fill

    def _generate_symbol_map(self, pattern: CrossStitchPattern) -> Dict[int, str]:
        """Generate symbol map for color indices with high-contrast characters."""
        # Professional cross-stitch symbols (high contrast, easily distinguishable)
        symbols = [
            "●", "○", "■", "□", "▲", "△", "♦", "♢", "★", "☆",
            "▼", "▽", "◆", "◇", "♠", "♣", "♥", "♪", "※", "⊕",
            "⊗", "⊙", "⊚", "⊛", "◉", "◎", "◯", "◐", "◑", "◒",
            "◓", "◔", "◕", "◖", "◗", "◘", "◙", "◚", "◛", "◜",
            "◝", "◞", "◟", "◠", "◡", "◢", "◣", "◤", "◥", "◦"
        ]

        # Create mapping for each unique color in the pattern
        symbol_map = {}
        unique_colors = set()

        # Collect unique color indices from pattern
        for y in range(pattern.height):
            for x in range(pattern.width):
                color_index = pattern.get_color_at(x, y)
                unique_colors.add(color_index)

        # Assign symbols to colors
        sorted_colors = sorted(list(unique_colors))
        for i, color_index in enumerate(sorted_colors):
            symbol_map[color_index] = symbols[i % len(symbols)]

        return symbol_map

    def _get_contrasting_font_color(self, background_color) -> str:
        """Calculate contrasting font color (black or white) for background color.

        Args:
            background_color: Color object with r, g, b values

        Returns:
            Hex color string ("000000" for black, "FFFFFF" for white)
        """
        # Calculate luminance using standard formula
        # Luminance = 0.299*R + 0.587*G + 0.114*B
        luminance = (
            0.299 * background_color.r
            + 0.587 * background_color.g
            + 0.114 * background_color.b
        ) / 255

        # Use white text on dark backgrounds, black text on light backgrounds
        return "FFFFFF" if luminance < 0.5 else "000000"

    def _apply_pattern_colors(
        self, worksheet: Worksheet, pattern: CrossStitchPattern
    ) -> None:
        """Apply background colors and symbols to cells based on pattern (with ruler offset)."""
        try:
            # Generate symbol map for this pattern
            symbol_map = self._generate_symbol_map(pattern)

            for y in range(pattern.height):
                for x in range(pattern.width):
                    # Get color for this position
                    color_index = pattern.get_color_at(x, y)
                    color = pattern.palette[color_index]

                    # Convert to Excel cell reference (offset by 1 for rulers)
                    cell = worksheet.cell(row=y + 2, column=x + 2)

                    # Create fill pattern with the color
                    fill_color = color.hex_code.lstrip("#")  # Remove # if present
                    fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

                    # Apply fill to cell
                    cell.fill = fill

                    # Add symbol instead of DMC code for better readability
                    symbol = symbol_map[color_index]

                    # Set cell format and data type to text
                    cell.number_format = "@"  # '@' is Excel's text format
                    cell.data_type = "s"  # Explicitly set as string type
                    cell.value = symbol

                    # Calculate contrasting font color for readability
                    font_color = self._get_contrasting_font_color(color)
                    cell.font = Font(color=font_color, size=12, bold=True)  # Larger font for symbols

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to apply pattern colors: {e}",
                sheet_name=pattern.resolution_name,
                cause=e,
            )

    def _apply_cell_styling(
        self, worksheet: Worksheet, pattern: CrossStitchPattern
    ) -> None:
        """Apply additional styling to pattern cells with Rule of 10 major grid lines."""
        try:
            # Define border styles for professional cross-stitch grid
            thin_border = Side(border_style="thin", color="000000")
            medium_border = Side(border_style="medium", color="000000")  # Major grid lines

            # Define center alignment
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Apply styling to all pattern cells with Rule of 10 implementation (offset for rulers)
            for y in range(1, pattern.height + 1):
                for x in range(1, pattern.width + 1):
                    cell = worksheet.cell(row=y + 1, column=x + 1)  # +1 offset for rulers

                    # Start with thin borders
                    top = thin_border
                    left = thin_border
                    right = thin_border
                    bottom = thin_border

                    # Rule of 10: Apply medium borders every 10th row/column
                    if x % 10 == 0:  # Every 10th column gets thick right border
                        right = medium_border
                    if y % 10 == 0:  # Every 10th row gets thick bottom border
                        bottom = medium_border

                    # Create border with appropriate thickness
                    border = Border(top=top, left=left, right=right, bottom=bottom)

                    # Apply border and alignment
                    cell.border = border
                    cell.alignment = center_alignment

            # Add professional freeze panes and print preparation
            self._setup_freeze_panes_and_print(worksheet, pattern)

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to apply cell styling: {e}",
                sheet_name=pattern.resolution_name,
                cause=e,
            )

    def _setup_freeze_panes_and_print(self, worksheet: Worksheet, pattern: CrossStitchPattern) -> None:
        """Set up freeze panes and print preparation for professional cross-stitch charts."""
        # Freeze panes at B2 (after rulers) so headers never disappear when scrolling
        worksheet.freeze_panes = "B2"

        # Print preparation: Ensure rulers appear on every printed page
        worksheet.print_title_rows = "1:1"  # Row 1 (column rulers) on every page
        worksheet.print_title_cols = "A:A"   # Column A (row rulers) on every page

        # Set print orientation to landscape for better cross-stitch chart viewing
        worksheet.page_setup.orientation = "landscape"

        # Set print scaling to fit more on page
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0  # Let height scale naturally

        # Add margins appropriate for cross-stitch charts
        worksheet.page_margins.left = 0.5
        worksheet.page_margins.right = 0.5
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

    def _create_legend_sheet(self, workbook: Workbook, pattern_set: PatternSet) -> None:
        """Create color legend sheet showing all colors used."""
        try:
            sheet_name = self._sanitize_sheet_name(self.config.legend_sheet_name)
            legend_sheet = workbook.create_sheet(title=sheet_name)

            # Set up legend headers (include Symbol and DMC information)
            headers = [
                "Symbol",
                "Color",
                "Hex Code",
                "RGB",
                "DMC Code",
                "Thread Name",
                "Usage Count",
                "Usage %",
            ]
            for col, header in enumerate(headers, 1):
                cell = legend_sheet.cell(row=1, column=col, value=header)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row = 2
            total_stitches = 0

            # Collect color usage statistics from all patterns
            color_usage = self._collect_color_usage_stats(pattern_set)

            # Calculate total stitches across all patterns
            for pattern in pattern_set.patterns.values():
                total_stitches += pattern.total_stitches

            # Add each color to the legend with symbols
            for color_data in color_usage:
                color = color_data["color"]
                count = color_data["count"]
                percentage = (count / total_stitches * 100) if total_stitches > 0 else 0

                # Get symbol for this color from first pattern that uses it
                symbol = "●"  # Default symbol
                for pattern in pattern_set.patterns.values():
                    symbol_map = self._generate_symbol_map(pattern)
                    # Find this color in the pattern
                    for color_idx, pattern_color in enumerate(pattern.palette):
                        if (pattern_color.hex_code == color.hex_code and
                            color_idx in symbol_map):
                            symbol = symbol_map[color_idx]
                            break
                    if symbol != "●":  # Found a symbol, break
                        break

                # Symbol column
                symbol_cell = legend_sheet.cell(row=row, column=1, value=symbol)
                symbol_cell.font = Font(size=16, bold=True)
                symbol_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Color sample cell (moved to column 2)
                color_cell = legend_sheet.cell(row=row, column=2)
                fill_color = color.hex_code.lstrip("#")
                color_cell.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
                color_cell.value = "■"  # Color block symbol
                color_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Hex code (column 3)
                legend_sheet.cell(row=row, column=3, value=color.hex_code)

                # RGB values (column 4)
                rgb_text = f"({color.r}, {color.g}, {color.b})"
                legend_sheet.cell(row=row, column=4, value=rgb_text)

                # DMC Code (column 5)
                dmc_cell = legend_sheet.cell(row=row, column=5)
                if color.thread_code:
                    dmc_cell.number_format = (
                        "@"  # Text format to prevent Excel warnings
                    )
                    dmc_cell.data_type = "s"  # Explicitly set as string type
                    dmc_cell.value = color.thread_code
                else:
                    dmc_cell.value = ""

                # Thread Name (column 6)
                legend_sheet.cell(
                    row=row, column=6, value=color.name if color.name else ""
                )

                # Usage count (column 7)
                legend_sheet.cell(row=row, column=7, value=count)

                # Usage percentage (column 8)
                legend_sheet.cell(row=row, column=8, value=f"{percentage:.1f}%")

                row += 1

            # Auto-adjust column widths
            self._auto_adjust_column_widths(legend_sheet)

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to create legend sheet: {e}",
                sheet_name=self.config.legend_sheet_name,
                cause=e,
            )

    def _create_summary_sheet(
        self, workbook: Workbook, pattern_set: PatternSet
    ) -> None:
        """Create summary sheet with pattern metadata."""
        try:
            summary_sheet = workbook.create_sheet(title="Summary")

            # Add metadata
            row = 1

            # Source image information
            summary_sheet.cell(row=row, column=1, value="Source Image:")
            summary_sheet.cell(
                row=row, column=2, value=str(pattern_set.source_image_path.name)
            )
            row += 1

            summary_sheet.cell(row=row, column=1, value="Generated On:")
            summary_sheet.cell(
                row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            row += 2

            # Pattern information
            summary_sheet.cell(row=row, column=1, value="Pattern Resolutions:")
            row += 1

            for resolution_name, pattern in pattern_set.patterns.items():
                summary_sheet.cell(row=row, column=1, value=f"  {resolution_name}:")
                summary_sheet.cell(
                    row=row,
                    column=2,
                    value=f"{pattern.width}x{pattern.height} stitches",
                )
                summary_sheet.cell(
                    row=row, column=3, value=f"{pattern.unique_colors_used} colors"
                )
                row += 1

            row += 1

            # Configuration information
            summary_sheet.cell(row=row, column=1, value="Configuration:")
            row += 1

            config_items = [
                ("Quantization Method", self.config.quantization_method),
                ("Max Colors", str(self.config.max_colors)),
                ("Transparency Handling", self.config.handle_transparency),
                ("Aspect Ratio Preserved", str(self.config.preserve_aspect_ratio)),
                ("Cell Size", f"{self.config.excel_cell_size} points"),
            ]

            for label, value in config_items:
                summary_sheet.cell(row=row, column=1, value=f"  {label}:")
                summary_sheet.cell(row=row, column=2, value=value)
                row += 1

            # Auto-adjust column widths
            self._auto_adjust_column_widths(summary_sheet)

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to create summary sheet: {e}", sheet_name="Summary", cause=e
            )

    def _collect_color_usage_stats(
        self, pattern_set: PatternSet
    ) -> List[Dict[str, Any]]:
        """Collect color usage statistics across all patterns."""
        color_counts = {}

        for pattern in pattern_set.patterns.values():
            usage_stats = pattern.get_color_usage_stats()

            for color_index, count in usage_stats.items():
                color = pattern.palette[color_index]
                color_key = color.rgb_tuple

                if color_key in color_counts:
                    color_counts[color_key]["count"] += count
                else:
                    color_counts[color_key] = {"color": color, "count": count}

        # Sort by usage count (most used first)
        return sorted(color_counts.values(), key=lambda x: x["count"], reverse=True)

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel compatibility."""
        # Excel sheet names can't contain: \ / * ? : [ ]
        invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
        sanitized = name

        for char in invalid_chars:
            sanitized = sanitized.replace(char, "_")

        # Limit length to 31 characters (Excel limit)
        if len(sanitized) > 31:
            sanitized = sanitized[:31]

        return sanitized

    def _auto_adjust_column_widths(self, worksheet: Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at reasonable width
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _save_workbook_to_bytes(self, workbook: Workbook) -> bytes:
        """Save workbook to bytes for file output."""
        try:
            # Create BytesIO object to capture Excel data
            excel_buffer = io.BytesIO()

            # Save workbook to buffer
            workbook.save(excel_buffer)

            # Get bytes data
            excel_data = excel_buffer.getvalue()
            excel_buffer.close()

            return excel_data

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to save workbook to bytes: {e}", cause=e
            )
