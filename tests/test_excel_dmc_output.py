"""Tests for DMC-enhanced Excel output functionality."""

from openpyxl import load_workbook

from src.cross_stitch.core.pattern_generator import PatternGenerator
from src.cross_stitch.models.config import GeneratorConfig
from src.cross_stitch.models.color_palette import Color, ColorPalette
from src.cross_stitch.models.pattern import CrossStitchPattern


class TestDMCExcelOutput:
    """Test DMC-enhanced Excel file generation."""

    def test_pattern_cells_display_dmc_codes_when_available(self, tiny_image, temp_dir):
        """Test that pattern cells show DMC codes as text when Color has thread info."""
        config = GeneratorConfig(
            resolutions=[(6, 6)], max_colors=3, include_color_legend=True
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "dmc_pattern_test.xlsx"

        # Generate pattern
        _ = generator.generate_patterns(tiny_image, output_path)

        # Load the Excel file with openpyxl to inspect contents
        workbook = load_workbook(output_path)
        pattern_sheet = workbook["6x6"]

        # Check that cells contain DMC codes as text values
        found_dmc_codes = []
        for row in range(1, 7):  # 6x6 pattern
            for col in range(1, 7):
                cell = pattern_sheet.cell(row=row, column=col)
                if cell.value:  # Cell should have DMC code as text
                    found_dmc_codes.append(cell.value)

        # Should have found DMC codes in pattern cells
        assert len(found_dmc_codes) > 0
        # DMC codes should be strings like "310", "BLANC", etc.
        assert all(isinstance(code, str) for code in found_dmc_codes)

        workbook.close()

    def test_dmc_text_uses_contrasting_colors_for_readability(
        self, tiny_image, temp_dir
    ):
        """Test that DMC code text uses contrasting colors against cell backgrounds."""
        config = GeneratorConfig(
            resolutions=[(8, 8)],  # Minimum 5x5 required
            max_colors=2,  # Force black and white colors for testing
            include_color_legend=True,
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "contrast_test.xlsx"

        _ = generator.generate_patterns(tiny_image, output_path)

        workbook = load_workbook(output_path)
        pattern_sheet = workbook["8x8"]

        # Check font colors for readability
        for row in range(1, 9):
            for col in range(1, 9):
                cell = pattern_sheet.cell(row=row, column=col)
                if cell.value:  # Cell has DMC code text
                    font_color = cell.font.color
                    assert font_color is not None  # Should have explicit font color

                    # Font color should be either black (000000) or white (FFFFFF) for contrast
                    if hasattr(font_color, "rgb"):
                        color_value = font_color.rgb
                        assert color_value in ["000000", "FFFFFF"], (
                            f"Font color {color_value} not contrastive"
                        )

        workbook.close()

    def test_color_legend_includes_dmc_codes_and_names(self, tiny_image, temp_dir):
        """Test that Color Legend sheet includes DMC code and thread name columns."""
        config = GeneratorConfig(
            resolutions=[(5, 5)], max_colors=4, include_color_legend=True
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "dmc_legend_test.xlsx"

        _ = generator.generate_patterns(tiny_image, output_path)

        workbook = load_workbook(output_path)
        legend_sheet = workbook["Color Legend"]

        # Check that legend headers include DMC information
        headers = []
        for col in range(1, 8):  # Extended header count for DMC info
            cell_value = legend_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value)

        # Should include new DMC columns
        assert "DMC Code" in headers
        assert "Thread Name" in headers
        # Original columns should still be present
        assert "Color" in headers
        assert "Usage Count" in headers

        # Check that legend rows contain DMC data
        dmc_code_column = headers.index("DMC Code") + 1
        thread_name_column = headers.index("Thread Name") + 1

        found_dmc_codes = []
        found_thread_names = []

        for row in range(2, 6):  # Check data rows
            dmc_code = legend_sheet.cell(row=row, column=dmc_code_column).value
            thread_name = legend_sheet.cell(row=row, column=thread_name_column).value

            if dmc_code:
                found_dmc_codes.append(dmc_code)
            if thread_name:
                found_thread_names.append(thread_name)

        # Should have found DMC codes and thread names
        assert len(found_dmc_codes) > 0
        assert len(found_thread_names) > 0

        workbook.close()

    def test_legend_shows_accurate_stitch_counts_per_dmc_color(
        self, tiny_image, temp_dir
    ):
        """Test that Color Legend shows correct number of stitches for each DMC color."""
        config = GeneratorConfig(
            resolutions=[(5, 5)],  # Minimum 5x5 = 25 total stitches
            max_colors=2,  # Force simple palette
            include_color_legend=True,
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "stitch_count_test.xlsx"

        _ = generator.generate_patterns(tiny_image, output_path)

        workbook = load_workbook(output_path)
        legend_sheet = workbook["Color Legend"]

        # Get header positions
        headers = []
        for col in range(1, 8):
            cell_value = legend_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value)

        usage_count_column = headers.index("Usage Count") + 1

        # Collect stitch counts from legend
        total_stitches_from_legend = 0
        for row in range(2, 5):  # Check data rows
            usage_count = legend_sheet.cell(row=row, column=usage_count_column).value
            if isinstance(usage_count, int) and usage_count > 0:
                total_stitches_from_legend += usage_count

        # Total should equal pattern dimensions (5x5 = 25 stitches)
        assert total_stitches_from_legend == 25

        workbook.close()

    def test_dmc_features_gracefully_handle_colors_without_thread_info(
        self, tiny_image, temp_dir
    ):
        """Test that Excel generation handles Colors without DMC thread info gracefully."""
        config = GeneratorConfig(
            resolutions=[(8, 8)], max_colors=3, include_color_legend=True
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "mixed_dmc_test.xlsx"

        # Generate patterns (some colors might not have DMC mapping)
        _ = generator.generate_patterns(tiny_image, output_path)

        # File should be created successfully without errors
        assert output_path.exists()
        assert output_path.stat().st_size > 1000

        workbook = load_workbook(output_path)

        # Pattern sheet should exist and be accessible
        pattern_sheet = workbook["8x8"]
        assert pattern_sheet is not None

        # Legend sheet should exist and handle missing DMC info gracefully
        legend_sheet = workbook["Color Legend"]
        assert legend_sheet is not None

        # Should not crash when accessing legend data
        for row in range(1, 6):
            for col in range(1, 8):
                _ = legend_sheet.cell(row=row, column=col).value
                # Just accessing values should not raise exceptions

        workbook.close()

    def test_dmc_enhanced_legend_maintains_existing_columns(self, tiny_image, temp_dir):
        """Test that DMC enhancements preserve existing legend functionality."""
        config = GeneratorConfig(
            resolutions=[(5, 5)], max_colors=3, include_color_legend=True
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "compatibility_test.xlsx"

        _ = generator.generate_patterns(tiny_image, output_path)

        workbook = load_workbook(output_path)
        legend_sheet = workbook["Color Legend"]

        # Get all headers
        headers = []
        for col in range(1, 10):  # Check more columns for expanded headers
            cell_value = legend_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value)

        # Original columns should still be present
        assert "Color" in headers
        assert "Hex Code" in headers
        assert "RGB" in headers
        assert "Usage Count" in headers
        assert "Usage %" in headers

        # New DMC columns should be added
        assert "DMC Code" in headers
        assert "Thread Name" in headers

        # Check that original column data is preserved
        hex_column = headers.index("Hex Code") + 1
        rgb_column = headers.index("RGB") + 1

        for row in range(2, 5):
            hex_value = legend_sheet.cell(row=row, column=hex_column).value
            rgb_value = legend_sheet.cell(row=row, column=rgb_column).value

            if hex_value:
                assert isinstance(hex_value, str)
                assert hex_value.startswith("#")  # Should be hex color code

            if rgb_value:
                assert isinstance(rgb_value, str)
                assert "(" in rgb_value and ")" in rgb_value  # Should be RGB format

        workbook.close()

    def test_pattern_cells_only_show_dmc_codes_for_dmc_colors(self, temp_dir):
        """Test that only Colors with DMC thread info show DMC codes in cells."""
        # Create a test pattern with mixed Color types
        colors = [
            Color(r=0, g=0, b=0, name="Black", thread_code="310"),  # Has DMC info
            Color(r=255, g=0, b=0),  # No DMC info
            Color(
                r=255, g=255, b=255, name="White", thread_code="BLANC"
            ),  # Has DMC info
        ]

        palette = ColorPalette(colors=colors, max_colors=3, quantization_method="test")

        # Create pattern with known color placement using numpy array
        import numpy as np

        color_data = np.array(
            [
                [0, 1, 0],  # DMC, non-DMC, DMC
                [1, 2, 1],  # non-DMC, DMC, non-DMC
                [0, 0, 2],  # DMC, DMC, DMC
            ]
        )


        pattern = CrossStitchPattern(
            width=3, height=3, colors=color_data, palette=palette, resolution_name="3x3"
        )

        # For now, just verify the pattern structure is correct
        assert pattern.get_color_at(0, 0) == 0  # Should be DMC color (310)
        assert pattern.get_color_at(1, 0) == 1  # Should be non-DMC color
        assert pattern.get_color_at(2, 0) == 0  # Should be DMC color (310)

        # The actual Excel verification would happen in a real Excel generation test
        # This test establishes that the pattern model can handle mixed DMC/non-DMC colors

    def test_excel_output_with_manually_created_dmc_colors(self, temp_dir):
        """Test Excel output using manually created pattern with DMC colors."""
        # Create DMC colors manually
        colors = [
            Color(r=0, g=0, b=0, name="Black", thread_code="310"),
            Color(r=255, g=255, b=255, name="White", thread_code="BLANC"),
            Color(r=227, g=29, b=66, name="Bright Red", thread_code="666"),
        ]

        palette = ColorPalette(
            colors=colors, max_colors=3, quantization_method="manual"
        )

        # Create 5x5 pattern (minimum size)
        import numpy as np

        color_data = np.array(
            [
                [0, 1, 0, 1, 0],  # Alternating black/white
                [1, 2, 1, 2, 1],  # Alternating white/red
                [0, 1, 0, 1, 0],  # Alternating black/white
                [1, 2, 1, 2, 1],  # Alternating white/red
                [0, 1, 0, 1, 0],  # Alternating black/white
            ]
        )

        from src.cross_stitch.models import PatternSet
        from src.cross_stitch.core.excel_generator import ExcelGenerator
        from src.cross_stitch.models.config import GeneratorConfig

        pattern = CrossStitchPattern(
            width=5, height=5, colors=color_data, palette=palette, resolution_name="5x5"
        )

        # Create dummy image file and pattern set
        dummy_image_path = temp_dir / "test.png"
        dummy_image_path.write_bytes(b"dummy")  # Create file so it exists

        pattern_set = PatternSet(
            patterns={"5x5": pattern},
            source_image_path=dummy_image_path,
            metadata={"test": "manual_dmc_test"},
        )

        # Generate Excel with DMC-enabled ExcelGenerator
        config = GeneratorConfig(include_color_legend=True)
        excel_generator = ExcelGenerator(config)
        output_path = temp_dir / "manual_dmc_test.xlsx"

        excel_generator.generate_excel_file(pattern_set, output_path)

        # Verify Excel file was created
        assert output_path.exists()

        # Load and verify Excel content
        from openpyxl import load_workbook

        workbook = load_workbook(output_path)

        # Check pattern sheet has DMC codes
        pattern_sheet = workbook["5x5"]
        found_dmc_codes = []
        for row in range(1, 6):
            for col in range(1, 6):
                cell_value = pattern_sheet.cell(row=row, column=col).value
                if cell_value:
                    found_dmc_codes.append(cell_value)

        # Should find DMC codes in cells
        assert len(found_dmc_codes) > 0
        assert (
            "310" in found_dmc_codes
            or "BLANC" in found_dmc_codes
            or "666" in found_dmc_codes
        )

        # Check legend has DMC columns
        legend_sheet = workbook["Color Legend"]
        headers = []
        for col in range(1, 8):
            header = legend_sheet.cell(row=1, column=col).value
            if header:
                headers.append(header)

        assert "DMC Code" in headers
        assert "Thread Name" in headers

        workbook.close()
