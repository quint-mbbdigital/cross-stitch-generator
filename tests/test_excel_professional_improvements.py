"""Test professional Excel improvements for cross-stitch patterns."""

import pytest
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image
import numpy as np

from src.cross_stitch.models.config import GeneratorConfig
from src.cross_stitch.core.pattern_generator import PatternGenerator


class TestExcelProfessionalImprovements:
    """Test professional cross-stitch Excel improvements."""

    @pytest.fixture
    def sample_config(self):
        """Create sample generator configuration."""
        return GeneratorConfig(
            resolutions=[(30, 30)],
            max_colors=8,
            quantization_method="median_cut",
            excel_cell_size=20.0,
            include_color_legend=True
        )

    @pytest.fixture
    def sample_image(self):
        """Create a simple test image."""
        # Create a 60x60 test image with distinct color blocks
        image_array = np.zeros((60, 60, 3), dtype=np.uint8)

        # Create color blocks for testing
        image_array[0:20, 0:20] = [255, 0, 0]     # Red
        image_array[0:20, 20:40] = [0, 255, 0]   # Green
        image_array[0:20, 40:60] = [0, 0, 255]   # Blue
        image_array[20:40, 0:20] = [255, 255, 0] # Yellow
        image_array[20:40, 20:40] = [255, 0, 255] # Magenta
        image_array[20:40, 40:60] = [0, 255, 255] # Cyan
        image_array[40:60, 0:30] = [128, 128, 128] # Gray
        image_array[40:60, 30:60] = [255, 255, 255] # White

        return Image.fromarray(image_array)

    def test_rule_of_10_major_grid_lines(self, sample_config, sample_image):
        """Test that Rule of 10 creates major grid lines every 10th row/column."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save sample image to temporary file
            image_path = Path(temp_dir) / "test_image.png"
            sample_image.save(image_path)

            output_path = Path(temp_dir) / "test_pattern.xlsx"

            # Generate pattern using correct API
            generator = PatternGenerator(sample_config)
            pattern_set = generator.generate_patterns(image_path, output_path)

            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            pattern_sheet = workbook.worksheets[0]  # First sheet should be pattern

            # Check that cells at 10th positions have medium borders
            # Account for ruler offset (+1)
            cell_10_10 = pattern_sheet.cell(row=11, column=11)  # 10th pattern cell
            cell_20_20 = pattern_sheet.cell(row=21, column=21)  # 20th pattern cell

            # These cells should have thicker borders on their right/bottom edges
            assert cell_10_10.border.right.style == "medium"
            assert cell_10_10.border.bottom.style == "medium"

            print("✅ Rule of 10 major grid lines working correctly")

    def test_custom_ruler_system(self, sample_config, sample_image):
        """Test that custom rulers show stitch counters."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save sample image to temporary file
            image_path = Path(temp_dir) / "test_image.png"
            sample_image.save(image_path)

            output_path = Path(temp_dir) / "test_pattern.xlsx"

            # Generate pattern using correct API
            generator = PatternGenerator(sample_config)
            pattern_set = generator.generate_patterns(image_path, output_path)

            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            pattern_sheet = workbook.worksheets[0]

            # Check column rulers (Row 1)
            assert pattern_sheet.cell(row=1, column=11).value == 10  # 10th column marker
            assert pattern_sheet.cell(row=1, column=21).value == 20  # 20th column marker

            # Check row rulers (Column A)
            assert pattern_sheet.cell(row=11, column=1).value == 10  # 10th row marker
            assert pattern_sheet.cell(row=21, column=1).value == 20  # 20th row marker

            # Check A1 is styled but empty
            assert pattern_sheet.cell(row=1, column=1).value is None
            assert pattern_sheet.cell(row=1, column=1).fill.start_color.index == "F0F0F0"

            print("✅ Custom ruler system working correctly")

    def test_symbols_instead_of_dmc_codes(self, sample_config, sample_image):
        """Test that cells contain symbols instead of DMC codes."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save sample image to temporary file
            image_path = Path(temp_dir) / "test_image.png"
            sample_image.save(image_path)

            output_path = Path(temp_dir) / "test_pattern.xlsx"

            # Generate pattern using correct API
            generator = PatternGenerator(sample_config)
            pattern_set = generator.generate_patterns(image_path, output_path)

            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            pattern_sheet = workbook.worksheets[0]

            # Check that pattern cells contain symbols (offset for rulers)
            cell_2_2 = pattern_sheet.cell(row=2, column=2)  # First pattern cell
            cell_value = cell_2_2.value

            # Should be a symbol, not a DMC code
            symbols = ["●", "○", "■", "□", "▲", "△", "♦", "♢", "★", "☆"]
            assert cell_value in symbols, f"Expected symbol, got: {cell_value}"

            # Check font size is larger for symbols
            assert cell_2_2.font.size == 12
            assert cell_2_2.font.bold is True

            print("✅ Symbol system working correctly")

    def test_freeze_panes_and_print_setup(self, sample_config, sample_image):
        """Test freeze panes and print preparation features."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save sample image to temporary file
            image_path = Path(temp_dir) / "test_image.png"
            sample_image.save(image_path)

            output_path = Path(temp_dir) / "test_pattern.xlsx"

            # Generate pattern using correct API
            generator = PatternGenerator(sample_config)
            pattern_set = generator.generate_patterns(image_path, output_path)

            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            pattern_sheet = workbook.worksheets[0]

            # Check freeze panes
            assert pattern_sheet.freeze_panes == "B2"

            # Check print settings
            assert pattern_sheet.print_title_rows == "1:1"
            assert pattern_sheet.print_title_cols == "A:A"
            assert pattern_sheet.page_setup.orientation == "landscape"
            assert pattern_sheet.page_setup.fitToWidth == 1
            assert pattern_sheet.page_setup.fitToHeight == 0

            print("✅ Freeze panes and print setup working correctly")

    def test_legend_includes_symbols(self, sample_config, sample_image):
        """Test that legend sheet includes symbol column."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save sample image to temporary file
            image_path = Path(temp_dir) / "test_image.png"
            sample_image.save(image_path)

            output_path = Path(temp_dir) / "test_pattern.xlsx"

            # Generate pattern using correct API
            generator = PatternGenerator(sample_config)
            pattern_set = generator.generate_patterns(image_path, output_path)

            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            legend_sheet = None

            # Find legend sheet
            for sheet in workbook.worksheets:
                if "legend" in sheet.title.lower():
                    legend_sheet = sheet
                    break

            assert legend_sheet is not None, "Legend sheet not found"

            # Check headers include Symbol
            headers = []
            for col in range(1, 9):
                header_cell = legend_sheet.cell(row=1, column=col)
                if header_cell.value:
                    headers.append(header_cell.value)

            expected_headers = ["Symbol", "Color", "Hex Code", "RGB", "DMC Code", "Thread Name", "Usage Count", "Usage %"]
            assert headers == expected_headers, f"Expected {expected_headers}, got {headers}"

            # Check that symbols appear in first data row
            symbol_cell = legend_sheet.cell(row=2, column=1)
            symbols = ["●", "○", "■", "□", "▲", "△", "♦", "♢", "★", "☆"]
            assert symbol_cell.value in symbols
            assert symbol_cell.font.size == 16
            assert symbol_cell.font.bold is True

            print("✅ Legend symbol column working correctly")

    def test_pattern_offset_for_rulers(self, sample_config, sample_image):
        """Test that pattern is properly offset for ruler system."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save sample image to temporary file
            image_path = Path(temp_dir) / "test_image.png"
            sample_image.save(image_path)

            output_path = Path(temp_dir) / "test_pattern.xlsx"

            # Generate pattern using correct API
            generator = PatternGenerator(sample_config)
            pattern_set = generator.generate_patterns(image_path, output_path)

            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            pattern_sheet = workbook.worksheets[0]

            # Check that row 1 and column A are rulers, pattern starts at B2
            assert pattern_sheet.cell(row=1, column=2).value is not None or pattern_sheet.cell(row=1, column=2).fill.start_color.index == "F0F0F0"  # Ruler
            assert pattern_sheet.cell(row=2, column=1).value is not None or pattern_sheet.cell(row=2, column=1).fill.start_color.index == "F0F0F0"  # Ruler

            # Pattern should start at B2
            pattern_cell = pattern_sheet.cell(row=2, column=2)
            assert pattern_cell.value is not None  # Should have symbol
            assert pattern_cell.fill.start_color.index != "F0F0F0"  # Should have color, not ruler gray

            print("✅ Pattern offset working correctly")


if __name__ == "__main__":
    # Quick manual test
    test_instance = TestExcelProfessionalImprovements()

    # Create test fixtures
    config = test_instance.sample_config()
    image = test_instance.sample_image()

    print("🧵 Testing Professional Excel Cross-Stitch Improvements")
    print("=" * 60)

    try:
        test_instance.test_rule_of_10_major_grid_lines(config, image)
        test_instance.test_custom_ruler_system(config, image)
        test_instance.test_symbols_instead_of_dmc_codes(config, image)
        test_instance.test_freeze_panes_and_print_setup(config, image)
        test_instance.test_legend_includes_symbols(config, image)
        test_instance.test_pattern_offset_for_rulers(config, image)

        print("\n🎉 All professional Excel improvements working perfectly!")
        print("✅ Rule of 10 major grid lines")
        print("✅ Custom ruler system with stitch counters")
        print("✅ Symbols instead of DMC codes for readability")
        print("✅ Freeze panes and print preparation")
        print("✅ Legend with symbol column")
        print("✅ Pattern offset for professional layout")

    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()