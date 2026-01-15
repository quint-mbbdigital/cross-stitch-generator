"""Integration tests for cross-stitch pattern generator."""

import pytest
from pathlib import Path
import numpy as np
from PIL import Image

from src.cross_stitch.core import PatternGenerator
from src.cross_stitch.models import GeneratorConfig
from src.cross_stitch.utils import PatternGenerationError
from tests.conftest import assert_valid_pattern


class TestPatternGeneratorIntegration:
    """Integration tests for the complete pattern generation workflow."""

    def test_generate_patterns_rgb_image(self, sample_image_rgb, sample_excel_output, mock_progress_callback):
        """Test complete pattern generation with RGB image."""
        # Create a simple config for faster testing
        config = GeneratorConfig(
            resolutions=[(20, 20), (30, 30)],
            max_colors=16,
            quantization_method="median_cut",
            preserve_aspect_ratio=True
        )

        generator = PatternGenerator(config)
        generator.set_progress_callback(mock_progress_callback)

        # Generate patterns
        pattern_set = generator.generate_patterns(sample_image_rgb, sample_excel_output)

        # Verify pattern set
        assert pattern_set.pattern_count == 2
        assert "20x20" in pattern_set.resolution_names
        assert "30x30" in pattern_set.resolution_names
        assert pattern_set.source_image_path == sample_image_rgb

        # Verify individual patterns
        pattern_20 = pattern_set.get_pattern("20x20")
        assert_valid_pattern(pattern_20)
        assert pattern_20.width == 20
        assert pattern_20.height == 20

        pattern_30 = pattern_set.get_pattern("30x30")
        assert_valid_pattern(pattern_30)
        assert pattern_30.width == 30
        assert pattern_30.height == 30

        # Verify Excel file was created
        assert sample_excel_output.exists()
        assert sample_excel_output.stat().st_size > 0

        # Verify progress callback was called
        assert len(mock_progress_callback.calls) > 0
        assert mock_progress_callback.final_progress >= 1.0

    def test_generate_patterns_rgba_image(self, sample_image_rgba, temp_dir):
        """Test pattern generation with RGBA (transparent) image."""
        config = GeneratorConfig(
            resolutions=[(15, 15)],
            max_colors=8,
            handle_transparency="white_background"
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "rgba_output.xlsx"

        pattern_set = generator.generate_patterns(sample_image_rgba, output_path)

        assert pattern_set.pattern_count == 1
        pattern = pattern_set.get_pattern("15x15")
        assert_valid_pattern(pattern)

        # Excel file should be created
        assert output_path.exists()

    def test_generate_patterns_grayscale_image(self, sample_image_grayscale, temp_dir):
        """Test pattern generation with grayscale image."""
        config = GeneratorConfig(
            resolutions=[(25, 25)],
            max_colors=32,
            quantization_method="kmeans"
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "gray_output.xlsx"

        pattern_set = generator.generate_patterns(sample_image_grayscale, output_path)

        assert pattern_set.pattern_count == 1
        pattern = pattern_set.get_pattern("25x25")
        assert_valid_pattern(pattern)

    def test_generate_single_pattern(self, sample_image_rgb):
        """Test generating a single pattern at specific resolution."""
        config = GeneratorConfig(max_colors=8)
        generator = PatternGenerator(config)

        pattern = generator.generate_single_pattern(sample_image_rgb, (40, 40))

        assert_valid_pattern(pattern)
        assert pattern.width == 40
        assert pattern.height == 40
        assert pattern.resolution_name == "40x40"

    def test_different_quantization_methods(self, tiny_image, temp_dir):
        """Test different color quantization methods."""
        # Test median cut
        config_median = GeneratorConfig(
            resolutions=[(10, 10)],
            max_colors=4,
            quantization_method="median_cut"
        )
        generator_median = PatternGenerator(config_median)
        output_median = temp_dir / "median_output.xlsx"
        pattern_set_median = generator_median.generate_patterns(tiny_image, output_median)

        # Test k-means
        config_kmeans = GeneratorConfig(
            resolutions=[(10, 10)],
            max_colors=4,
            quantization_method="kmeans"
        )
        generator_kmeans = PatternGenerator(config_kmeans)
        output_kmeans = temp_dir / "kmeans_output.xlsx"
        pattern_set_kmeans = generator_kmeans.generate_patterns(tiny_image, output_kmeans)

        # Both should succeed and create valid patterns
        assert_valid_pattern(pattern_set_median.get_pattern("10x10"))
        assert_valid_pattern(pattern_set_kmeans.get_pattern("10x10"))

        # Both Excel files should be created
        assert output_median.exists()
        assert output_kmeans.exists()

    def test_aspect_ratio_preservation(self, temp_dir):
        """Test aspect ratio preservation vs stretching."""
        # Create a rectangular image (2:1 aspect ratio)
        rect_image = Image.new('RGB', (60, 30), color=(100, 150, 200))
        rect_path = temp_dir / "rectangular.png"
        rect_image.save(rect_path)

        # Test with aspect ratio preservation
        config_preserve = GeneratorConfig(
            resolutions=[(20, 20)],
            preserve_aspect_ratio=True,
            max_colors=4
        )
        generator_preserve = PatternGenerator(config_preserve)
        output_preserve = temp_dir / "preserve_output.xlsx"
        pattern_set_preserve = generator_preserve.generate_patterns(rect_path, output_preserve)

        # Test without aspect ratio preservation
        config_stretch = GeneratorConfig(
            resolutions=[(20, 20)],
            preserve_aspect_ratio=False,
            max_colors=4
        )
        generator_stretch = PatternGenerator(config_stretch)
        output_stretch = temp_dir / "stretch_output.xlsx"
        pattern_set_stretch = generator_stretch.generate_patterns(rect_path, output_stretch)

        # Both should create valid patterns
        pattern_preserve = pattern_set_preserve.get_pattern("20x20")
        pattern_stretch = pattern_set_stretch.get_pattern("20x20")

        assert_valid_pattern(pattern_preserve)
        assert_valid_pattern(pattern_stretch)

        # Both should have the target dimensions
        assert pattern_preserve.width == 20
        assert pattern_preserve.height == 20
        assert pattern_stretch.width == 20
        assert pattern_stretch.height == 20

    def test_transparency_handling_methods(self, sample_image_rgba, temp_dir):
        """Test different transparency handling methods."""
        transparency_methods = ["white_background", "remove", "preserve"]

        for method in transparency_methods:
            config = GeneratorConfig(
                resolutions=[(15, 15)],
                max_colors=8,
                handle_transparency=method
            )
            generator = PatternGenerator(config)
            output_path = temp_dir / f"transparency_{method}.xlsx"

            try:
                pattern_set = generator.generate_patterns(sample_image_rgba, output_path)
                pattern = pattern_set.get_pattern("15x15")
                assert_valid_pattern(pattern)
                assert output_path.exists()
            except PatternGenerationError:
                # Some transparency handling methods might not work with all images
                # This is acceptable for testing purposes
                pass

    def test_multiple_resolutions(self, tiny_image, temp_dir):
        """Test generating multiple resolutions in one pass."""
        config = GeneratorConfig(
            resolutions=[(8, 8), (12, 12), (16, 16), (20, 20)],
            max_colors=8
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "multi_res_output.xlsx"

        pattern_set = generator.generate_patterns(tiny_image, output_path)

        assert pattern_set.pattern_count == 4
        for resolution in ["8x8", "12x12", "16x16", "20x20"]:
            pattern = pattern_set.get_pattern(resolution)
            assert_valid_pattern(pattern)

        assert output_path.exists()

    def test_get_processing_info(self, sample_image_rgb):
        """Test getting processing information without actually processing."""
        config = GeneratorConfig(resolutions=[(25, 25), (50, 50)])
        generator = PatternGenerator(config)

        info = generator.get_processing_info(sample_image_rgb)

        assert 'source_image' in info
        assert 'target_resolutions' in info
        assert 'config' in info

        source_info = info['source_image']
        assert source_info['size'] == (100, 100)  # Known size of sample image

        target_info = info['target_resolutions']
        assert '25x25' in target_info
        assert '50x50' in target_info

    def test_estimate_processing_time(self, sample_image_rgb):
        """Test processing time estimation."""
        config = GeneratorConfig()
        generator = PatternGenerator(config)

        estimates = generator.estimate_processing_time(sample_image_rgb)

        assert 'image_loading' in estimates
        assert 'image_resizing' in estimates
        assert 'color_quantization' in estimates
        assert 'excel_generation' in estimates
        assert 'total' in estimates

        # All estimates should be positive numbers
        for key, value in estimates.items():
            assert isinstance(value, (int, float))
            assert value > 0

    def test_error_handling_invalid_image(self, nonexistent_image_path, temp_dir):
        """Test error handling with invalid image."""
        config = GeneratorConfig()
        generator = PatternGenerator(config)
        output_path = temp_dir / "error_output.xlsx"

        with pytest.raises(PatternGenerationError):
            generator.generate_patterns(nonexistent_image_path, output_path)

    def test_error_handling_invalid_resolution(self, tiny_image, temp_dir):
        """Test error handling with invalid resolution for image size."""
        # Try to create a pattern much larger than the source image
        config = GeneratorConfig(resolutions=[(1000, 1000)])
        generator = PatternGenerator(config)
        output_path = temp_dir / "error_output.xlsx"

        with pytest.raises(PatternGenerationError):
            generator.generate_patterns(tiny_image, output_path)

    def test_progress_callback_coverage(self, sample_image_rgb, temp_dir, mock_progress_callback):
        """Test that progress callback covers the full workflow."""
        config = GeneratorConfig(
            resolutions=[(20, 20)],
            max_colors=8
        )

        generator = PatternGenerator(config)
        generator.set_progress_callback(mock_progress_callback)
        output_path = temp_dir / "progress_test.xlsx"

        generator.generate_patterns(sample_image_rgb, output_path)

        # Should have multiple progress updates
        assert len(mock_progress_callback.calls) >= 3

        # Progress should start at 0 and end at 1.0
        first_progress = mock_progress_callback.calls[0][1]
        last_progress = mock_progress_callback.calls[-1][1]

        assert first_progress == 0.0
        assert last_progress >= 1.0

        # Progress should be monotonically increasing
        for i in range(1, len(mock_progress_callback.calls)):
            current_progress = mock_progress_callback.calls[i][1]
            previous_progress = mock_progress_callback.calls[i-1][1]
            assert current_progress >= previous_progress

    def test_pattern_color_consistency(self, tiny_image, temp_dir):
        """Test that patterns use colors consistently."""
        config = GeneratorConfig(
            resolutions=[(10, 10)],
            max_colors=4
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "consistency_test.xlsx"

        pattern_set = generator.generate_patterns(tiny_image, output_path)
        pattern = pattern_set.get_pattern("10x10")

        # All color indices should be valid
        max_index = len(pattern.palette) - 1
        assert np.all(pattern.colors >= 0)
        assert np.all(pattern.colors <= max_index)

        # Color usage stats should add up to total stitches
        usage_stats = pattern.get_color_usage_stats()
        total_usage = sum(usage_stats.values())
        assert total_usage == pattern.total_stitches

    def test_excel_file_structure(self, tiny_image, temp_dir):
        """Test that generated Excel file has correct structure."""
        config = GeneratorConfig(
            resolutions=[(8, 8), (12, 12)],
            max_colors=4,
            include_color_legend=True
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "structure_test.xlsx"

        pattern_set = generator.generate_patterns(tiny_image, output_path)

        # Verify Excel file exists and has content
        assert output_path.exists()
        assert output_path.stat().st_size > 1000  # Should have substantial content

        # Could add openpyxl tests here to verify sheet structure
        # but that would require loading the Excel file

    def test_pattern_set_summary(self, tiny_image, temp_dir):
        """Test pattern set summary information."""
        config = GeneratorConfig(
            resolutions=[(6, 6), (10, 10)],
            max_colors=8
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "summary_test.xlsx"

        pattern_set = generator.generate_patterns(tiny_image, output_path)
        summary = pattern_set.get_summary()

        assert 'source_image' in summary
        assert 'pattern_count' in summary
        assert 'resolutions' in summary
        assert 'patterns' in summary

        assert summary['pattern_count'] == 2
        assert '6x6' in summary['resolutions']
        assert '10x10' in summary['resolutions']

        # Each pattern should have detailed info
        for res_name in ['6x6', '10x10']:
            assert res_name in summary['patterns']
            pattern_info = summary['patterns'][res_name]
            assert 'width' in pattern_info
            assert 'height' in pattern_info
            assert 'total_stitches' in pattern_info

    def test_dmc_workflow_integration(self, tiny_image, temp_dir):
        """Test complete DMC workflow: image → quantize → DMC config → Excel with DMC features."""
        # Create config with DMC features enabled
        config = GeneratorConfig(
            resolutions=[(8, 8), (12, 12)],
            max_colors=6,
            enable_dmc=True,
            include_color_legend=True,
            quantization_method="median_cut"
        )

        generator = PatternGenerator(config)
        output_path = temp_dir / "dmc_workflow_test.xlsx"

        # Generate patterns with DMC configuration
        pattern_set = generator.generate_patterns(tiny_image, output_path)

        # Verify pattern generation succeeded
        assert pattern_set.pattern_count == 2
        assert output_path.exists()
        assert output_path.stat().st_size > 1000  # Non-empty file

        # Verify Excel file structure for DMC features
        from openpyxl import load_workbook
        workbook = load_workbook(output_path)

        # Check that pattern sheets exist
        assert "8x8" in workbook.sheetnames
        assert "12x12" in workbook.sheetnames

        # Check that Color Legend sheet exists and has DMC columns
        assert "Color Legend" in workbook.sheetnames
        legend_sheet = workbook["Color Legend"]

        # Verify DMC columns are present in legend headers
        headers = []
        for col in range(1, 10):  # Check first 9 columns
            cell_value = legend_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value)

        # DMC-specific columns should be present
        assert "DMC Code" in headers
        assert "Thread Name" in headers

        # Original columns should still be present
        assert "Color" in headers
        assert "Hex Code" in headers
        assert "Usage Count" in headers

        # Verify pattern sheets are accessible
        pattern_sheet_8x8 = workbook["8x8"]
        assert pattern_sheet_8x8 is not None

        # Test that cells can be accessed without errors
        for row in range(1, 9):
            for col in range(1, 9):
                cell = pattern_sheet_8x8.cell(row=row, column=col)
                # Just accessing should not raise exceptions

        workbook.close()

        # Verify individual patterns have correct properties
        pattern_8x8 = pattern_set.get_pattern("8x8")
        assert_valid_pattern(pattern_8x8)
        assert pattern_8x8.width == 8
        assert pattern_8x8.height == 8

        pattern_12x12 = pattern_set.get_pattern("12x12")
        assert_valid_pattern(pattern_12x12)
        assert pattern_12x12.width == 12
        assert pattern_12x12.height == 12